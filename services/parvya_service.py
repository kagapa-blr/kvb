import logging
import csv
import os
from io import StringIO, BytesIO
from logging import getLogger

from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from flask import send_file, current_app

from model.models import (
    db,
    Parva,
    Sandhi,
    Padya,
    GamakaVachana,
)
from services.gamaka_vachana_service import GamakaVachanaService
from utils.path_builder import GamakaPathBuilder

logger = getLogger()


# -------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------

def normalize_file_path(file_path):
    """
    Normalize file paths - convert absolute paths to relative paths.
    Uses GamakaPathBuilder utility for consistent path handling.
    """
    return GamakaPathBuilder.normalize_path(file_path)


def delete_gamaka_file(file_path, file_type="photo"):
    """
    Delete a gamaka photo or audio file from the filesystem.
    
    Args:
        file_path: Relative path like 'photos/gamakaPhotos/1_1_1_file.jpg'
        file_type: 'photo' or 'audio' for logging
    
    Returns:
        True if deleted successfully, False if not found or error
    """
    if not file_path:
        return False
    
    try:
        from flask import current_app
        if current_app and current_app.static_folder:
            # Convert relative path to absolute filesystem path
            full_path = os.path.join(current_app.static_folder, file_path)
            
            if os.path.isfile(full_path):
                os.remove(full_path)
                logger.info(f"Deleted gamaka {file_type}: {full_path}")
                return True
            else:
                logger.debug(f"Gamaka {file_type} not found at: {full_path}")
                return False
    except Exception as e:
        logger.error(f"Error deleting gamaka {file_type} at {file_path}: {e}")
        return False


def rename_gamaka_file_with_metadata(old_path, parva_number, sandhi_number, padya_number, gamaka_name, raga, file_type="photo"):
    """
    Rename a gamaka file to include updated name and raga in the filename.
    
    Args:
        old_path: Current relative path like 'photos/gamakaPhotos/1_1_1_old.jpg'
        parva_number, sandhi_number, padya_number: Context numbers
        gamaka_name: New gamaka_vachakara_name
        raga: New raga
        file_type: 'photo' or 'audio'
    
    Returns:
        New relative path if successful, None if operation failed
    """
    if not old_path or not gamaka_name or not raga:
        return None
    
    try:
        from flask import current_app
        if not current_app or not current_app.static_folder:
            return None
        
        # Get file extension
        _, ext = os.path.splitext(old_path)
        if not ext:
            return None
        
        # Build new filename
        if file_type == "photo":
            new_filename = f"{int(parva_number)}_{int(sandhi_number)}_{int(padya_number)}_{gamaka_name}_{raga}{ext}"
            new_path = f"{GamakaPathBuilder.PHOTOS_SUBDIR}/{new_filename}"
        else:  # audio
            new_filename = f"{int(parva_number)}-{int(sandhi_number)}-{int(padya_number)}-{gamaka_name}-{raga}{ext}"
            new_path = f"{GamakaPathBuilder.AUDIO_SUBDIR}/{new_filename}"
        
        # Get full filesystem paths
        old_full_path = os.path.join(current_app.static_folder, old_path)
        new_full_path = os.path.join(current_app.static_folder, new_path)
        
        # Check if old file exists
        if not os.path.isfile(old_full_path):
            logger.debug(f"Old gamaka {file_type} not found at: {old_full_path}")
            return None
        
        # Avoid overwriting if new file already exists
        if os.path.isfile(new_full_path):
            logger.warning(f"Target gamaka {file_type} already exists at: {new_full_path}")
            # Remove old file to avoid confusion
            try:
                os.remove(old_full_path)
                logger.info(f"Deleted old gamaka {file_type} due to existing target: {old_full_path}")
            except Exception as e:
                logger.error(f"Error deleting old gamaka {file_type}: {e}")
            return new_path
        
        # Rename the file
        os.rename(old_full_path, new_full_path)
        logger.info(f"Renamed gamaka {file_type} from {old_full_path} to {new_full_path}")
        
        return new_path
    
    except Exception as e:
        logger.error(f"Error renaming gamaka {file_type} from {old_path}: {e}")
        return None


def get_gamaka_photo_path_with_fs_check(gamaka_vachana_obj, parva_number, sandhi_number, padya_number):
    """
    Get gamaka photo path with filesystem fallback check.
    
    Logic:
    1. Check if DB has a valid photo path
    2. If empty/None, search filesystem based on parva/sandhi/padya numbers
    3. Return first match found, or None if nothing found
    """
    if not gamaka_vachana_obj:
        return None
    
    # Check DB first
    if gamaka_vachana_obj.gamaka_vachakar_photo_path:
        normalized = GamakaPathBuilder.normalize_path(gamaka_vachana_obj.gamaka_vachakar_photo_path)
        if normalized:
            return normalized
    
    # If DB is empty, try filesystem
    try:
        from flask import current_app
        if current_app:
            static_folder = current_app.static_folder
            photos_dir = os.path.join(static_folder, 'photos', 'gamakaPhotos')
            
            if os.path.isdir(photos_dir):
                matches = GamakaPathBuilder.find_photos(photos_dir, parva_number, sandhi_number, padya_number)
                if matches:
                    # Return first match found (already sorted by GamakaPathBuilder)
                    relative_path = GamakaPathBuilder.construct_relative_photo_path(
                        parva_number, sandhi_number, padya_number, matches[0]
                    )
                    return relative_path
    except Exception as e:
        logger.debug(f"Error during photo filesystem check: {e}")
    
    return None


def get_gamaka_audio_path_with_fs_check(gamaka_vachana_obj, parva_number, sandhi_number, padya_number):
    """
    Get gamaka audio path with filesystem fallback check.
    
    Logic:
    1. Check if DB has a valid audio path
    2. If empty/None, search filesystem based on parva/sandhi/padya numbers
    3. Return first match found, or None if nothing found
    """
    if not gamaka_vachana_obj:
        return None
    
    # Check DB first
    if gamaka_vachana_obj.gamaka_vachakar_audio_path:
        normalized = GamakaPathBuilder.normalize_path(gamaka_vachana_obj.gamaka_vachakar_audio_path)
        if normalized:
            return normalized
    
    # If DB is empty, try filesystem
    try:
        from flask import current_app
        if current_app:
            static_folder = current_app.static_folder
            audio_dir = os.path.join(static_folder, 'audio', 'gamakaAudio')
            
            if os.path.isdir(audio_dir):
                matches = GamakaPathBuilder.find_audios(audio_dir, parva_number, sandhi_number, padya_number)
                if matches:
                    # Return first match found (already sorted by GamakaPathBuilder)
                    relative_path = GamakaPathBuilder.construct_relative_audio_path(
                        parva_number, sandhi_number, padya_number, matches[0]
                    )
                    return relative_path
    except Exception as e:
        logger.debug(f"Error during audio filesystem check: {e}")
    
    return None



# -------------------------------------------------------
# COMMON UTILITIES
# -------------------------------------------------------

def commit_session():
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error("Database commit failed: %s", e)
        raise


def get_pagination(offset: int = 0, limit: int = 20):
    offset = max(offset, 0)
    limit = max(min(limit, 100), 1)
    # offset already provided directly
    return offset, limit


# -------------------------------------------------------
# PARVA SERVICE
# -------------------------------------------------------

class ParvaService:

    def get_all(self, offset=0, limit=20, **kwargs):
        try:
            offset, limit = get_pagination(offset, limit)

            query = Parva.query.order_by(Parva.parva_number)
            total = query.count()

            records = (
                query.offset(offset)
                .limit(limit)
                .all()
            )

            return {
                "offset": offset,
                "limit": limit,
                "total": total,
                "data": [
                    {
                        "id": p.id,
                        "name": p.name,
                        "parva_number": p.parva_number,
                        "sandhi_count": len(p.sandhis),
                        "parvantya": p.parvantya,
                    }
                    for p in records
                ],
            }, 200

        except Exception as e:
            logger.error("Error fetching parvas: %s", e)
            return {"error": "Database error"}, 500

    def get_by_number(self, parva_number, **kwargs):
        parva = Parva.query.filter_by(parva_number=parva_number).first()

        if not parva:
            return {"error": "Parva not found"}, 404

        return {
            "id": parva.id,
            "name": parva.name,
            "parva_number": parva.parva_number,
            "parvantya": parva.parvantya,
            "sandhi_count": len(parva.sandhis),
        }, 200

    def create(self, **kwargs):
        name = kwargs.get("name")

        if not name:
            return {"error": "Name required"}, 400

        try:
            max_number = db.session.query(
                func.max(Parva.parva_number)
            ).scalar()

            next_number = (max_number or 0) + 1

            record = Parva(
                name=name,
                parva_number=next_number,
                parvantya=kwargs.get("parvantya"),
            )

            db.session.add(record)
            commit_session()

            return {
                "id": record.id,
                "parva_number": record.parva_number,
                "name": record.name,
            }, 201

        except IntegrityError:
            return {"error": "Duplicate parva"}, 409

        except Exception as e:
            logger.error("Create parva failed: %s", e)
            return {"error": "Create failed"}, 500

    def update(self, parva_number, **kwargs):
        parva = Parva.query.filter_by(
            parva_number=parva_number
        ).first()

        if not parva:
            return {"error": "Parva not found"}, 404

        try:
            parva.name = kwargs.get("name", parva.name)
            parva.parvantya = kwargs.get(
                "parvantya", parva.parvantya
            )

            commit_session()

            return {"message": "Updated"}, 200

        except Exception as e:
            logger.error("Update parva failed: %s", e)
            return {"error": "Update failed"}, 500

    def delete(self, parva_number, **kwargs):
        parva = Parva.query.filter_by(
            parva_number=parva_number
        ).first()

        if not parva:
            return {"error": "Parva not found"}, 404

        try:
            db.session.delete(parva)
            commit_session()

            return {"message": "Deleted"}, 200

        except Exception as e:
            logger.error("Delete parva failed: %s", e)
            return {"error": "Delete failed"}, 500

    def search(self, query="", offset=0, limit=20, **kwargs):
        """
        Search parva by name or number.
        Handles both numeric (number) and text (name) queries.
        """
        try:
            offset, limit = get_pagination(offset, limit)
            
            # Check if query is numeric
            is_numeric = False
            search_number = None
            try:
                search_number = int(query)
                is_numeric = True
            except (ValueError, TypeError):
                pass
            
            # Build query
            base_query = Parva.query.order_by(Parva.parva_number)
            
            if is_numeric and search_number is not None:
                # Search by parva number (exact match)
                base_query = base_query.filter(Parva.parva_number == search_number)
            elif query:
                # Search by name (case-insensitive partial match)
                base_query = base_query.filter(
                    Parva.name.ilike(f"%{query}%")
                )
            
            total = base_query.count()
            records = base_query.offset(offset).limit(limit).all()
            
            return {
                "offset": offset,
                "limit": limit,
                "total": total,
                "data": [
                    {
                        "id": p.id,
                        "name": p.name,
                        "parva_number": p.parva_number,
                        "sandhi_count": len(p.sandhis),
                        "parvantya": p.parvantya,
                    }
                    for p in records
                ],
            }, 200
            
        except Exception as e:
            logger.error("Parva search failed: %s", e)
            return {"error": "Search failed"}, 500


# -------------------------------------------------------
# SANDHI SERVICE
# -------------------------------------------------------

class SandhiService:

    def get_by_parva(
        self,
        parva_number,
        offset=0,
        limit=20,
        **kwargs,
    ):
        offset, limit = get_pagination(offset, limit)

        parva = Parva.query.filter_by(
            parva_number=parva_number
        ).first()

        if not parva:
            return {"error": "Parva not found"}, 404

        query = (
            Sandhi.query.options(joinedload(Sandhi.parva))
            .filter_by(parva_id=parva.id)
            .order_by(Sandhi.sandhi_number)
        )

        total = query.count()

        records = (
            query.offset(offset)
            .limit(limit)
            .all()
        )

        return {
            "offset": offset,
            "limit": limit,
            "total": total,
            "data": [
                {
                    "id": s.id,
                    "parva_number": s.parva.parva_number,
                    "sandhi_number": s.sandhi_number,
                    "name": s.name,
                    "padya_count": len(s.padyas),
                    "padya_numbers": sorted([p.padya_number for p in s.padyas]),
                }
                for s in records
            ],
        }, 200

    def get_unique(
        self,
        parva_number,
        sandhi_number,
        **kwargs,
    ):
        sandhi = (
            Sandhi.query.join(Parva)
            .filter(
                Parva.parva_number == parva_number,
                Sandhi.sandhi_number == sandhi_number,
            )
            .first()
        )

        if not sandhi:
            return {"error": "Sandhi not found"}, 404

        return {
            "id": sandhi.id,
            "parva_number": parva_number,
            "sandhi_number": sandhi.sandhi_number,
            "name": sandhi.name,
            "padya_count": len(sandhi.padyas),
            "padya_numbers": sorted([p.padya_number for p in sandhi.padyas]),
        }, 200

    def create(self, **kwargs):
        parva_number = kwargs.get("parva_number")
        name = kwargs.get("name")

        if not parva_number:
            return {"error": "Parva number required"}, 400

        if not name:
            return {"error": "Name required"}, 400

        parva = Parva.query.filter_by(
            parva_number=parva_number
        ).first()

        if not parva:
            return {"error": "Parva not found"}, 404

        try:
            max_number = db.session.query(
                func.max(Sandhi.sandhi_number)
            ).filter_by(parva_id=parva.id).scalar()

            next_number = (max_number or 0) + 1

            record = Sandhi(
                parva_id=parva.id,
                name=name,
                sandhi_number=next_number,
            )

            db.session.add(record)
            commit_session()

            return {
                "parva_number": parva_number,
                "sandhi_number": record.sandhi_number,
            }, 201

        except IntegrityError:
            return {"error": "Duplicate sandhi"}, 409

        except Exception as e:
            logger.error("Create sandhi failed: %s", e)
            return {"error": "Create failed"}, 500

    def update(
        self,
        parva_number,
        sandhi_number,
        **kwargs,
    ):
        sandhi = (
            Sandhi.query.join(Parva)
            .filter(
                Parva.parva_number == parva_number,
                Sandhi.sandhi_number == sandhi_number,
            )
            .first()
        )

        if not sandhi:
            return {"error": "Sandhi not found"}, 404

        try:
            sandhi.name = kwargs.get(
                "name", sandhi.name
            )

            commit_session()

            return {"message": "Updated"}, 200

        except Exception as e:
            logger.error("Update sandhi failed: %s", e)
            return {"error": "Update failed"}, 500

    def delete(
        self,
        parva_number,
        sandhi_number,
        **kwargs,
    ):
        sandhi = (
            Sandhi.query.join(Parva)
            .filter(
                Parva.parva_number == parva_number,
                Sandhi.sandhi_number == sandhi_number,
            )
            .first()
        )

        if not sandhi:
            return {"error": "Sandhi not found"}, 404

        try:
            db.session.delete(sandhi)
            commit_session()

            return {"message": "Deleted"}, 200

        except Exception as e:
            logger.error("Delete sandhi failed: %s", e)
            return {"error": "Delete failed"}, 500

    def search(self, parva_number, query="", offset=0, limit=20, **kwargs):
        """
        Search sandhi by name or number within a specific parva.
        Handles both numeric (number) and text (name) queries.
        """
        try:
            offset, limit = get_pagination(offset, limit)
            
            # Check if parva exists
            parva = Parva.query.filter_by(parva_number=parva_number).first()
            if not parva:
                return {"error": "Parva not found"}, 404
            
            # Check if query is numeric
            is_numeric = False
            search_number = None
            try:
                search_number = int(query)
                is_numeric = True
            except (ValueError, TypeError):
                pass
            
            # Build query
            base_query = (
                Sandhi.query.options(joinedload(Sandhi.parva))
                .filter_by(parva_id=parva.id)
                .order_by(Sandhi.sandhi_number)
            )
            
            if is_numeric and search_number is not None:
                # Search by sandhi number (exact match)
                base_query = base_query.filter(Sandhi.sandhi_number == search_number)
            elif query:
                # Search by name (case-insensitive partial match)
                base_query = base_query.filter(Sandhi.name.ilike(f"%{query}%"))
            
            total = base_query.count()
            records = base_query.offset(offset).limit(limit).all()
            
            return {
                "offset": offset,
                "limit": limit,
                "total": total,
                "data": [
                    {
                        "parva_number": s.parva.parva_number,
                        "sandhi_number": s.sandhi_number,
                        "name": s.name,
                        "padya_count": len(s.padyas),
                    }
                    for s in records
                ],
            }, 200
            
        except Exception as e:
            logger.error("Sandhi search failed: %s", e)
            return {"error": "Search failed"}, 500


# -------------------------------------------------------
# PADYA SERVICE
# -------------------------------------------------------

class PadyaService:

    # ---------------------------------------------
    # SEARCH WITH PAGINATION
    # ---------------------------------------------

    def search(
        self,
        parva_number=None,
        sandhi_number=None,
        keyword=None,
        offset=0,
        limit=20,
        **kwargs,
    ):
        try:
            offset, limit = get_pagination(offset, limit)

            query = (
                Padya.query
                .join(Sandhi)
                .join(Parva)
            )

            if parva_number:
                query = query.filter(
                    Parva.parva_number == parva_number
                )

            if sandhi_number:
                query = query.filter(
                    Sandhi.sandhi_number == sandhi_number
                )

            if keyword:
                like = f"%{keyword}%"
                query = query.filter(
                    Padya.padya.ilike(like)
                )

            query = query.order_by(
                Padya.padya_number
            )

            total = query.count()

            records = (
                query.offset(offset)
                .limit(limit)
                .all()
            )

            return {
                "offset": offset,
                "limit": limit,
                "total": total,
                "data": [
                    {
                        "parva_name": r.sandhi.parva.name,
                        "parva_number": r.sandhi.parva.parva_number,
                        "sandhi_number": r.sandhi.sandhi_number,
                        "padya_number": r.padya_number,
                        "padya": r.padya,
                        "pathantar": r.pathantar,
                        "gadya": r.gadya,
                        "tippani": r.tippani,
                        "artha": r.artha,
                        "preview": (
                            r.padya[:80] + "..."
                            if r.padya
                            else None
                        ),
                    }
                    for r in records
                ],
            }, 200

        except Exception as e:
            logger.error("Search padya failed: %s", e)
            return {"error": "Search failed"}, 500

    # ---------------------------------------------
    # GET PADYA NUMBERS BY SANDHI (FOR DROPDOWNS)
    # ---------------------------------------------

    def get_numbers_by_sandhi(self, sandhi_id, **kwargs):
        """
        Optimized endpoint for getting just padya numbers for a sandhi.

        Returns only padya_number for efficient dropdown population.
        No pagination - returns ALL numbers for the sandhi.

        Args:
            sandhi_id: ID of the sandhi

        Returns:
            {
              "sandhi_id": sandhi_id,
              "padya_numbers": [1, 2, 3, ..., n]
            }
        """
        try:
            # Get sandhi to verify it exists
            sandhi = Sandhi.query.filter_by(id=sandhi_id).first()
            if not sandhi:
                return {"error": "Sandhi not found"}, 404

            # Get all padya numbers for this sandhi, sorted
            padyas = (
                Padya.query
                .filter_by(sandhi_id=sandhi_id)
                .order_by(Padya.padya_number.asc())
                .all()
            )

            padya_numbers = [p.padya_number for p in padyas]

            return {
                "sandhi_id": sandhi_id,
                "sandhi_number": sandhi.sandhi_number,
                "parva_number": sandhi.parva.parva_number,
                "padya_numbers": padya_numbers,
                "total": len(padya_numbers),
            }, 200

        except Exception as e:
            logger.error("Get padya numbers by sandhi failed: %s", e)
            return {"error": "Failed to fetch padya numbers"}, 500

    # ---------------------------------------------
    # FETCH UNIQUE PADYA
    # ---------------------------------------------

    def get_unique(
        self,
        parva_number,
        sandhi_number,
        padya_number,
        author_name=None,
        raga=None,
        **kwargs,
    ):
        record = (
            Padya.query.join(Sandhi)
            .join(Parva)
            .filter(
                Parva.parva_number == parva_number,
                Sandhi.sandhi_number == sandhi_number,
                Padya.padya_number == padya_number,
            )
            .first()
        )

        if not record:
            return {"error": "Padya not found"}, 404

        # Get sandhi and parva objects for names
        sandhi = record.sandhi
        parva = sandhi.parva if sandhi else None

        # Fetch GamakaVachana data for this padya
        gamaka_query = GamakaVachana.query.filter_by(
            parva_number=parva_number,
            sandhi_number=sandhi_number,
            padya_number=padya_number
        )

        # Filter by author name and raga if provided
        if author_name and author_name.strip():
            gamaka_query = gamaka_query.filter_by(gamaka_vachakara_name=author_name.strip())
        if raga and raga.strip():
            gamaka_query = gamaka_query.filter_by(raga=raga.strip())

        gamaka_vachana_list = gamaka_query.all()

        gamaka_vachana_data = []
        
        if gamaka_vachana_list:
            # Case 1: DB has records - use DB-first, FS-fallback for paths
            for gv in gamaka_vachana_list:
                # Get photo path using same logic as audio:
                # 1. Use DB value if present
                # 2. Search filesystem if DB is empty
                # 3. Return None if found nowhere
                photo_path = get_gamaka_photo_path_with_fs_check(
                    gv, parva_number, sandhi_number, padya_number
                )

                # Get audio path with same filesystem fallback logic
                audio_path = get_gamaka_audio_path_with_fs_check(
                    gv, parva_number, sandhi_number, padya_number
                )

                gamaka_vachana_data.append({
                    "id": gv.id,
                    "gamaka_vachakara_name": gv.gamaka_vachakara_name,
                    "raga": gv.raga,
                    "gamaka_vachakar_photo_path": photo_path,
                    "gamaka_vachakar_audio_path": audio_path,
                })
        else:
            # Case 2: No DB records - search filesystem directly
            try:
                from flask import current_app
                if current_app:
                    static_folder = current_app.static_folder
                    if static_folder:
                        # Search for photos
                        photos_dir = os.path.join(static_folder, 'photos', 'gamakaPhotos')
                        photo_matches = []
                        if os.path.isdir(photos_dir):
                            photo_matches = GamakaPathBuilder.find_photos(
                                photos_dir, parva_number, sandhi_number, padya_number
                            )

                        # Search for audios
                        audio_dir = os.path.join(static_folder, 'audio', 'gamakaAudio')
                        audio_matches = []
                        if os.path.isdir(audio_dir):
                            audio_matches = GamakaPathBuilder.find_audios(
                                audio_dir, parva_number, sandhi_number, padya_number
                            )

                        # If we found any files, create a "virtual" gamaka entry
                        if photo_matches or audio_matches:
                            photo_path = None
                            audio_path = None

                            if photo_matches:
                                photo_path = GamakaPathBuilder.construct_relative_photo_path(
                                    parva_number, sandhi_number, padya_number, photo_matches[0]
                                )

                            if audio_matches:
                                audio_path = GamakaPathBuilder.construct_relative_audio_path(
                                    parva_number, sandhi_number, padya_number, audio_matches[0]
                                )

                            gamaka_vachana_data.append({
                                "id": None,
                                "gamaka_vachakara_name": None,
                                "raga": None,
                                "gamaka_vachakar_photo_path": photo_path,
                                "gamaka_vachakar_audio_path": audio_path,
                            })
            except Exception as e:
                logger.debug(f"Error searching filesystem for gamaka files: {e}")



        return {
            "id": record.id,
            "parva_number": parva_number,
            "parva_name": parva.name if parva else "",
            "sandhi_number": sandhi_number,
            "sandhi_name": sandhi.name if sandhi else "",
            "padya_number": record.padya_number,
            "padya": record.padya,
            "artha": record.artha,
            "tippani": record.tippani,
            "gadya": record.gadya,
            "suchane": record.suchane,
            "pathantar": record.pathantar,
            "created": record.created.isoformat() if record.created and hasattr(record.created, 'isoformat') else record.created,
            "updated": record.updated.isoformat() if record.updated and hasattr(record.updated, 'isoformat') else record.updated,
            "updated_by": record.updated_by,
            "gamaka_vachana": gamaka_vachana_data,
        }, 200

    # ---------------------------------------------
    # CREATE
    # ---------------------------------------------

    def create(self, **kwargs):
        try:
            parva_number = kwargs.get(
                "parva_number"
            )
            sandhi_number = kwargs.get(
                "sandhi_number"
            )

            padya_text = kwargs.get("padya")

            if not (
                parva_number
                and sandhi_number
                and padya_text
            ):
                return {"error": "Required fields missing"}, 400

            sandhi = (
                Sandhi.query.join(Parva)
                .filter(
                    Parva.parva_number
                    == parva_number,
                    Sandhi.sandhi_number
                    == sandhi_number,
                )
                .first()
            )

            if not sandhi:
                return {"error": "Sandhi not found"}, 404

            max_number = db.session.query(
                func.max(Padya.padya_number)
            ).filter_by(
                sandhi_id=sandhi.id
            ).scalar()

            next_number = (max_number or 0) + 1

            record = Padya(
                sandhi_id=sandhi.id,
                padya_number=next_number,
                padya=padya_text,
                artha=kwargs.get("artha"),
                tippani=kwargs.get("tippani"),
                gadya=kwargs.get("gadya"),
                suchane=kwargs.get("suchane"),
                pathantar=kwargs.get("pathantar"),
                updated_by=kwargs.get("updated_by"),  # Store creator/initial editor
            )

            db.session.add(record)
            db.session.flush()  # Flush to get the ID

            # Handle GamakaVachana data if provided
            gamaka_vachakara_name = kwargs.get("gamaka_vachakara_name", "").strip()
            raga = kwargs.get("gamaka_raga", "").strip()
            gamaka_vachakar_photo_path = normalize_file_path(kwargs.get("gamaka_photo_path", "").strip())
            gamaka_vachakar_audio_path = normalize_file_path(kwargs.get("gamaka_audio_path", "").strip())

            # Only create if BOTH raga and gamaka_vachakara_name are provided (both are NOT NULL in DB)
            if gamaka_vachakara_name and raga:
                GamakaVachanaService.create(
                    parva_number=sandhi.parva.parva_number,
                    sandhi_number=sandhi.sandhi_number,
                    padya_number=next_number,
                    raga=raga,
                    gamaka_vachakara_name=gamaka_vachakara_name,
                    gamaka_vachakar_photo_path=gamaka_vachakar_photo_path,
                    gamaka_vachakar_audio_path=gamaka_vachakar_audio_path
                )

            commit_session()

            return {
                "parva_number": parva_number,
                "sandhi_number": sandhi_number,
                "padya_number": record.padya_number,
            }, 201

        except IntegrityError:
            return {"error": "Duplicate padya"}, 409

        except Exception as e:
            logger.error("Create padya failed: %s", e)
            return {"error": "Create failed"}, 500

    # ---------------------------------------------
    # UPDATE
    # ---------------------------------------------

    def update(
        self,
        parva_number,
        sandhi_number,
        padya_number,
        **kwargs,
    ):
        record = (
            Padya.query.join(Sandhi)
            .join(Parva)
            .filter(
                Parva.parva_number == parva_number,
                Sandhi.sandhi_number == sandhi_number,
                Padya.padya_number == padya_number,
            )
            .first()
        )

        if not record:
            return {"error": "Padya not found"}, 404

        try:
            record.padya = kwargs.get(
                "padya", record.padya
            )
            record.artha = kwargs.get(
                "artha", record.artha
            )
            record.tippani = kwargs.get(
                "tippani", record.tippani
            )
            record.gadya = kwargs.get(
                "gadya", record.gadya
            )
            record.suchane = kwargs.get(
                "suchane", record.suchane
            )
            record.pathantar = kwargs.get(
                "pathantar", record.pathantar
            )

            # Update the editor info if provided
            if "updated_by" in kwargs and kwargs["updated_by"]:
                record.updated_by = kwargs.get("updated_by")

            # Handle GamakaVachana data update
            sandhi = record.sandhi
            parva = sandhi.parva if sandhi else None

            # Get the raw values from kwargs first (before stripping)
            gamaka_vachakara_name_raw = kwargs.get("gamaka_vachakara_name")
            gamaka_raga_raw = kwargs.get("gamaka_raga")
            gamaka_photo_path_raw = kwargs.get("gamaka_photo_path")
            gamaka_audio_path_raw = kwargs.get("gamaka_audio_path")

            # Process and strip the values
            gamaka_vachakara_name = (gamaka_vachakara_name_raw or "").strip() if gamaka_vachakara_name_raw is not None else None
            raga = (gamaka_raga_raw or "").strip() if gamaka_raga_raw is not None else None
            gamaka_vachakar_photo_path = normalize_file_path(gamaka_photo_path_raw) if gamaka_photo_path_raw is not None else None
            gamaka_vachakar_audio_path = normalize_file_path(gamaka_audio_path_raw) if gamaka_audio_path_raw is not None else None

            # Get existing GamakaVachana for this padya
            existing_gamaka = GamakaVachana.query.filter_by(
                parva_number=parva.parva_number if parva else None,
                sandhi_number=sandhi.sandhi_number if sandhi else None,
                padya_number=padya_number
            ).first()

            # Check if there's any actual user-provided gamaka metadata (not just file paths)
            # Only create DB records if user provided BOTH raga AND gamaka_vachakara_name (both required in DB)
            has_user_gamaka_metadata = gamaka_vachakara_name and raga

            if existing_gamaka:
                # Track what changed to determine if we need to rename files
                name_changed = gamaka_vachakara_name and gamaka_vachakara_name != existing_gamaka.gamaka_vachakara_name
                raga_changed = raga and raga != existing_gamaka.raga
                
                # If name or raga is changing and we have existing files,
                # rename those files to reflect the new metadata
                if (name_changed or raga_changed):
                    # Use the new values, or fall back to existing values if not provided
                    final_name = gamaka_vachakara_name or existing_gamaka.gamaka_vachakara_name
                    final_raga = raga or existing_gamaka.raga
                    
                    # Rename photo file if it exists
                    if existing_gamaka.gamaka_vachakar_photo_path and final_name and final_raga:
                        new_photo_path = rename_gamaka_file_with_metadata(
                            existing_gamaka.gamaka_vachakar_photo_path,
                            parva.parva_number if parva else None,
                            sandhi.sandhi_number if sandhi else None,
                            padya_number,
                            final_name,
                            final_raga,
                            file_type="photo"
                        )
                        if new_photo_path:
                            gamaka_vachakar_photo_path = new_photo_path
                    
                    # Rename audio file if it exists
                    if existing_gamaka.gamaka_vachakar_audio_path and final_name and final_raga:
                        new_audio_path = rename_gamaka_file_with_metadata(
                            existing_gamaka.gamaka_vachakar_audio_path,
                            parva.parva_number if parva else None,
                            sandhi.sandhi_number if sandhi else None,
                            padya_number,
                            final_name,
                            final_raga,
                            file_type="audio"
                        )
                        if new_audio_path:
                            gamaka_vachakar_audio_path = new_audio_path
                
                # Update existing record - only update if a value is explicitly provided
                if gamaka_vachakara_name:
                    existing_gamaka.gamaka_vachakara_name = gamaka_vachakara_name
                if raga:
                    existing_gamaka.raga = raga

                # IMPORTANT: Check if key was SENT in request (not just if value is not None)
                # This handles: empty string "", null, and actual paths
                # Frontend sends "" or null when user deletes media - we must update DB to null
                if "gamaka_photo_path" in kwargs:
                    # Key was explicitly sent - update it (even if null or empty)
                    existing_gamaka.gamaka_vachakar_photo_path = gamaka_vachakar_photo_path

                if "gamaka_audio_path" in kwargs:
                    # Key was explicitly sent - update it (even if null or empty)
                    # This ensures deletion (null) actually gets saved to database
                    existing_gamaka.gamaka_vachakar_audio_path = gamaka_vachakar_audio_path
            elif has_user_gamaka_metadata:
                # Create new GamakaVachana ONLY if user provided BOTH raga and gamaka_vachakara_name
                # Do NOT create records just because files exist on filesystem
                GamakaVachanaService.create(
                    parva_number=parva.parva_number if parva else None,
                    sandhi_number=sandhi.sandhi_number if sandhi else None,
                    padya_number=padya_number,
                    raga=raga,
                    gamaka_vachakara_name=gamaka_vachakara_name,
                    gamaka_vachakar_photo_path=gamaka_vachakar_photo_path,
                    gamaka_vachakar_audio_path=gamaka_vachakar_audio_path
                )

            commit_session()

            return {"message": "Updated"}, 200

        except Exception as e:
            logger.error("Update padya failed: %s", e)
            return {"error": "Update failed"}, 500

    # ---------------------------------------------
    # BULK OPERATIONS
    # ---------------------------------------------

    def generate_template(self, **kwargs):
        """
        Generate a CSV template for bulk padya upload.

        Template columns:
        parva_number, parva_name, sandhi_number, sandhi_name, padya_number, padya, artha, tippani, gadya, suchane, pathantar, raga, gamaka_vachanakara_name

        Note: parva_name and sandhi_name are optional - used only if parva/sandhi needs to be created
        Note: raga and gamaka_vachanakara_name are optional - used to create gamaka_vachana records
        """
        try:
            from io import StringIO
            from flask import send_file
            import csv

            # Create CSV in memory
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                "parva_number",
                "parva_name",
                "sandhi_number",
                "sandhi_name",
                "padya_number",
                "padya",
                "artha",
                "tippani",
                "gadya",
                "suchane",
                "pathantar",
                "raga",
                "gamaka_vachanakara_name"
            ])

            # Write sample rows
            writer.writerow([
                "1",
                "ಆದಿ ಪರ್ವ",
                "1",
                "ಸಂಭವ ಸಂಧಿ",
                "",
                "ಪದ್ಯದ ಮೂಲ ಪಠ್ಯ",
                "ಪದ್ಯದ ಅರ್ಥ",
                "ಟಿಪ್ಪಣಿ",
                "ಗದ್ಯ",
                "ಸುಚನೆ",
                "ಪಠಾಂತರ",
                "ಧೀರಶಂಕರಾಭರಣ",
                "ಗಾಯಕರ ಹೆಸರು"
            ])
            writer.writerow([
                "1",
                "ಆದಿ ಪರ್ವ",
                "1",
                "ಸಂಭವ ಸಂಧಿ",
                "",
                "ಎರಡನೇ ಪದ್ಯ",
                "ಅರ್ಥ",
                "ಟಿಪ್ಪಣಿ",
                "ಗದ್ಯ",
                "ಸುಚನೆ",
                "ಪಠಾಂತರ",
                "ಭೈರವ",
                "ಇತರ ಗಾಯಕರು"
            ])

            # Get the CSV content
            csv_content = output.getvalue()
            output.close()

            # Return as file download
            from io import BytesIO
            file_buffer = BytesIO(csv_content.encode('utf-8-sig'))

            return send_file(
                file_buffer,
                mimetype='text/csv',
                as_attachment=True,
                download_name='padya_template.csv'
            )

        except Exception as e:
            logger.error("Template generation failed: %s", e)
            return {"error": "Template generation failed"}, 500

    def bulk_upload(self, file, **kwargs):
        """
        Bulk upload padya from CSV or Excel file.

        Expected columns:
        parva_number, parva_name, sandhi_number, sandhi_name, padya_number, padya, artha, tippani, gadya, suchane, pathantar, raga, gamaka_vachanakara_name

        Features:
        - Auto-creates Parva if parva_number and parva_name are provided
        - Auto-creates Sandhi if sandhi_number and sandhi_name are provided
        - Updates existing padya or creates new one if not found
        - Creates/updates GamakaVachana records if raga and gamaka_vachanakara_name are provided
        """
        try:
            import csv
            from io import StringIO, BytesIO

            filename = file.filename.lower()
            records_created = 0
            records_updated = 0
            records_failed = 0
            parvas_created = 0
            sandhis_created = 0
            errors = []

            if filename.endswith('.csv'):
                # Handle CSV
                stream = StringIO(file.stream.read().decode('utf-8'), newline=None)
                reader = csv.DictReader(stream)
                rows = list(reader)
            elif filename.endswith(('.xls', '.xlsx')):
                # Handle Excel
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(BytesIO(file.stream.read()))
                    ws = wb.active

                    # Read header
                    headers = [cell.value for cell in ws[1]]
                    rows = []
                    for row_idx in range(2, ws.max_row + 1):
                        row_dict = {}
                        for col_idx, header in enumerate(headers, 1):
                            row_dict[header] = ws.cell(row_idx, col_idx).value
                        rows.append(row_dict)
                except ImportError:
                    return {"error": "openpyxl not installed for Excel support"}, 400
            else:
                return {"error": "Unsupported file format. Use CSV or Excel."}, 400

            # Validate CSV structure FIRST
            if not rows:
                return {"error": "CSV file is empty. Please add data rows."}, 400

            # Clean headers: remove BOM and strip whitespace
            cleaned_rows = []
            for row in rows:
                cleaned_row = {}
                for key, value in row.items():
                    # Remove BOM from header keys
                    clean_key = key.encode('utf-8').lstrip(b'\xef\xbb\xbf').decode('utf-8').strip() if key else ''
                    cleaned_row[clean_key] = value
                cleaned_rows.append(cleaned_row)
            rows = cleaned_rows

            # Get headers from first row keys
            if rows:
                headers = list(rows[0].keys())
                # Check for required columns (case-insensitive)
                headers_lower = [h.lower() if h else '' for h in headers]
                required_cols = ['parva_number', 'sandhi_number', 'padya']
                missing_cols = []

                for req_col in required_cols:
                    if not any(h == req_col for h in headers_lower):
                        missing_cols.append(req_col)

                if missing_cols:
                    return {
                        "error": f"Missing required columns: {', '.join(missing_cols)}",
                        "given_columns": headers,
                        "required_columns": required_cols,
                        "missing_columns": missing_cols
                    }, 400

            # Validate each row has required field values
            validation_errors = []
            for row_idx, row in enumerate(rows, 2):
                parva_number_val = (row.get('parva_number') or '').strip()
                sandhi_number_val = (row.get('sandhi_number') or '').strip()
                padya_text = (row.get('padya') or '').strip()

                if not parva_number_val:
                    validation_errors.append(f"Row {row_idx}: 'parva_number' is empty (required)")
                if not sandhi_number_val:
                    validation_errors.append(f"Row {row_idx}: 'sandhi_number' is empty (required)")
                if not padya_text:
                    validation_errors.append(f"Row {row_idx}: 'padya' is empty (required)")

            if validation_errors:
                return {
                    "error": "Validation failed",
                    "validation_errors": validation_errors[:20],
                    "total_validation_errors": len(validation_errors),
                    "message": "Please fix the above errors before uploading"
                }, 400

            # Process each row
            for row_idx, row in enumerate(rows, 2):
                try:
                    # Get and strip values first (handle empty cells and whitespace)
                    parva_number_val = (row.get('parva_number') or '').strip()
                    parva_name = (row.get('parva_name') or '').strip()
                    sandhi_number_val = (row.get('sandhi_number') or '').strip()
                    sandhi_name = (row.get('sandhi_name') or '').strip()
                    padya_number_val = (row.get('padya_number') or '').strip()
                    padya_text = (row.get('padya') or '').strip()

                    # All values should be present at this point due to validation
                    # Convert to integers
                    try:
                        parva_number = int(parva_number_val)
                        sandhi_number = int(sandhi_number_val)
                    except ValueError:
                        errors.append(f"Row {row_idx}: parva_number or sandhi_number must be valid numbers")
                        records_failed += 1
                        continue

                    # padya_number is optional, parse if provided
                    padya_number = None
                    if padya_number_val:
                        try:
                            padya_number = int(padya_number_val)
                        except ValueError:
                            errors.append(f"Row {row_idx}: padya_number must be a valid number")
                            records_failed += 1
                            continue

                    # Find or create Parva
                    parva = Parva.query.filter_by(parva_number=parva_number).first()
                    if not parva:
                        if not parva_name:
                            errors.append(f"Row {row_idx}: Parva {parva_number} not found and parva_name not provided")
                            records_failed += 1
                            continue
                        try:
                            # Auto-create Parva
                            parva = Parva(
                                parva_number=parva_number,
                                name=parva_name
                            )
                            db.session.add(parva)
                            db.session.flush()  # Flush to get the ID
                            parvas_created += 1
                        except IntegrityError:
                            db.session.rollback()
                            parva = Parva.query.filter_by(parva_number=parva_number).first()
                            if not parva:
                                errors.append(f"Row {row_idx}: Failed to create or find Parva {parva_number}")
                                records_failed += 1
                                continue

                    # Find or create Sandhi
                    sandhi = Sandhi.query.filter_by(
                        parva_id=parva.id,
                        sandhi_number=sandhi_number
                    ).first()

                    if not sandhi:
                        if not sandhi_name:
                            errors.append(f"Row {row_idx}: Sandhi {sandhi_number} not found in Parva {parva_number} and sandhi_name not provided")
                            records_failed += 1
                            continue
                        try:
                            # Auto-create Sandhi
                            sandhi = Sandhi(
                                parva_id=parva.id,
                                sandhi_number=sandhi_number,
                                name=sandhi_name
                            )
                            db.session.add(sandhi)
                            db.session.flush()  # Flush to get the ID
                            sandhis_created += 1
                        except IntegrityError:
                            db.session.rollback()
                            sandhi = Sandhi.query.filter_by(
                                parva_id=parva.id,
                                sandhi_number=sandhi_number
                            ).first()
                            if not sandhi:
                                errors.append(f"Row {row_idx}: Failed to create or find Sandhi {sandhi_number}")
                                records_failed += 1
                                continue

                    # Auto-generate padya_number if not provided
                    if not padya_number:
                        max_number = db.session.query(
                            func.max(Padya.padya_number)
                        ).filter_by(sandhi_id=sandhi.id).scalar()
                        padya_number = (max_number or 0) + 1

                    # Check if padya already exists
                    existing = Padya.query.filter_by(
                        sandhi_id=sandhi.id,
                        padya_number=padya_number
                    ).first()

                    if existing:
                        # Update existing
                        existing.padya = padya_text
                        existing.artha = (row.get('artha') or '').strip() or None
                        existing.tippani = (row.get('tippani') or '').strip() or None
                        existing.gadya = (row.get('gadya') or '').strip() or None
                        existing.suchane = (row.get('suchane') or '').strip() or None
                        existing.pathantar = (row.get('pathantar') or '').strip() or None
                        records_updated += 1

                        # Handle gamaka_vachana details if provided
                        raga = (row.get('raga') or '').strip()
                        gamaka_vachakara_name = (row.get('gamaka_vachanakara_name') or '').strip()

                        if raga or gamaka_vachakara_name:
                            # Both fields should be present if either is provided
                            if not raga or not gamaka_vachakara_name:
                                errors.append(f"Row {row_idx}: If providing gamaka details, both raga and gamaka_vachanakara_name are required")
                            else:
                                # Check if gamaka_vachana already exists using service
                                gamakas = GamakaVachanaService.get_by_padya(
                                    parva.parva_number,
                                    sandhi.sandhi_number,
                                    padya_number
                                )
                                
                                gamaka = None
                                for g in gamakas:
                                    if g.gamaka_vachakara_name == gamaka_vachakara_name:
                                        gamaka = g
                                        break

                                if gamaka:
                                    # Update existing gamaka_vachana
                                    gamaka.raga = raga
                                else:
                                    # Create new gamaka_vachana using service
                                    GamakaVachanaService.create(
                                        parva_number=parva.parva_number,
                                        sandhi_number=sandhi.sandhi_number,
                                        padya_number=padya_number,
                                        raga=raga,
                                        gamaka_vachakara_name=gamaka_vachakara_name
                                    )
                    else:
                        # Create new
                        new_padya = Padya(
                            sandhi_id=sandhi.id,
                            padya_number=padya_number,
                            padya=padya_text,
                            artha=(row.get('artha') or '').strip() or None,
                            tippani=(row.get('tippani') or '').strip() or None,
                            gadya=(row.get('gadya') or '').strip() or None,
                            suchane=(row.get('suchane') or '').strip() or None,
                            pathantar=(row.get('pathantar') or '').strip() or None,
                        )
                        db.session.add(new_padya)
                        records_created += 1

                        # Handle gamaka_vachana details if provided
                        raga = (row.get('raga') or '').strip()
                        gamaka_vachakara_name = (row.get('gamaka_vachanakara_name') or '').strip()

                        if raga or gamaka_vachakara_name:
                            # Both fields should be present if either is provided
                            if not raga or not gamaka_vachakara_name:
                                errors.append(f"Row {row_idx}: If providing gamaka details, both raga and gamaka_vachanakara_name are required")
                            else:
                                # Create gamaka_vachana record using service
                                GamakaVachanaService.create(
                                    parva_number=parva.parva_number,
                                    sandhi_number=sandhi.sandhi_number,
                                    padya_number=padya_number,
                                    raga=raga,
                                    gamaka_vachakara_name=gamaka_vachakara_name
                                )

                except ValueError as ve:
                    errors.append(f"Row {row_idx}: Invalid number format - {str(ve)}")
                    records_failed += 1
                except IntegrityError as ie:
                    db.session.rollback()
                    errors.append(f"Row {row_idx}: Duplicate or constraint violation - {str(ie)}")
                    records_failed += 1
                except Exception as e:
                    db.session.rollback()
                    errors.append(f"Row {row_idx}: {str(e)}")
                    records_failed += 1

            # Commit all changes
            try:
                commit_session()
            except Exception as e:
                logger.error("Bulk upload commit failed: %s", e)
                return {"error": f"Failed to save data: {str(e)}"}, 500

            return {
                "message": "Bulk upload completed",
                "records_created": records_created,
                "records_updated": records_updated,
                "records_failed": records_failed,
                "parvas_created": parvas_created,
                "sandhis_created": sandhis_created,
                "errors": errors[:20],  # Return first 20 errors
                "total_errors": len(errors)
            }, 200

        except Exception as e:
            logger.error("Bulk upload failed: %s", e)
            return {"error": f"Bulk upload failed: {str(e)}"}, 500

    # ---------------------------------------------
    # DELETE
    # ---------------------------------------------

    def delete(
        self,
        parva_number,
        sandhi_number,
        padya_number,
        **kwargs,
    ):
        record = (
            Padya.query.join(Sandhi)
            .join(Parva)
            .filter(
                Parva.parva_number == parva_number,
                Sandhi.sandhi_number == sandhi_number,
                Padya.padya_number == padya_number,
            )
            .first()
        )

        if not record:
            return {"error": "Padya not found"}, 404

        try:
            db.session.delete(record)
            commit_session()

            return {"message": "Deleted"}, 200

        except Exception as e:
            logger.error("Delete padya failed: %s", e)
            return {"error": "Delete failed"}, 500

    # ---------------------------------------------
    # EXPORT
    # ---------------------------------------------

    def export_all(self, **kwargs):
        """
        Export all padyas with complete details as CSV.

        Columns: parva_number, parva_name, sandhi_number, sandhi_name,
                 padya_number, padya, artha, tippani, gadya, suchane, pathantar, raga, gamaka_vachanakara_name
        """
        try:
            from datetime import datetime
            from io import StringIO, BytesIO
            from flask import send_file
            import csv
            from model.models import GamakaVachana

            # Query all padyas with their parva and sandhi details
            query = (
                Padya.query
                .join(Sandhi)
                .join(Parva)
                .order_by(Parva.parva_number, Sandhi.sandhi_number, Padya.padya_number)
                .all()
            )

            if not query:
                return {"error": "No padyas found to export"}, 404

            # Create CSV in memory
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                "parva_number",
                "parva_name",
                "sandhi_number",
                "sandhi_name",
                "padya_number",
                "padya",
                "artha",
                "tippani",
                "gadya",
                "suchane",
                "pathantar",
                "raga",
                "gamaka_vachanakara_name"
            ])

            # Write all padya records
            for padya in query:
                sandhi = padya.sandhi
                parva = sandhi.parva if sandhi else None

                # Get gamaka details if they exist
                gamaka = GamakaVachana.query.filter_by(
                    parva_number=parva.parva_number if parva else None,
                    sandhi_number=sandhi.sandhi_number if sandhi else None,
                    padya_number=padya.padya_number
                ).first()

                writer.writerow([
                    parva.parva_number if parva else "",
                    parva.name if parva else "",
                    sandhi.sandhi_number if sandhi else "",
                    sandhi.name if sandhi else "",
                    padya.padya_number or "",
                    padya.padya or "",
                    padya.artha or "",
                    padya.tippani or "",
                    padya.gadya or "",
                    padya.suchane or "",
                    padya.pathantar or "",
                    gamaka.raga if gamaka else "",
                    gamaka.gamaka_vachakara_name if gamaka else ""
                ])

            # Get the CSV content
            csv_content = output.getvalue()
            output.close()

            # Generate filename with date
            current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"padya_export_{current_date}.csv"

            # Return as file download
            file_buffer = BytesIO(csv_content.encode('utf-8-sig'))

            return send_file(
                file_buffer,
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )

        except Exception as e:
            logger.error("Export padyas failed: %s", e)
            return {"error": f"Export failed: {str(e)}"}, 500

    # ---------------------------------------------
    # PHOTO UPLOAD
    # ---------------------------------------------

    def save_gamaka_photo(self, file, parva_number, sandhi_number, padya_number, raga, author_name, **kwargs):
        """
        Save uploaded photo for GamakaVachana (who sang the padya).

        Naming convention: parva_number_sandhi_number_padya_number_raga_authorname.ext
        Example: 1_1_1_Dheeravati_RamanandaSagara.jpg

        Returns the relative photo path for storage in database.
        """
        try:
            import os
            from werkzeug.utils import secure_filename

            if not file or file.filename == "":
                return {"error": "No file provided"}, 400

            # Validate file type - only JPEG and WebP
            allowed_extensions = {'jpg', 'jpeg', 'webp'}
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

            if file_ext not in allowed_extensions:
                return {"error": "ಫೋಟೋ ಸ್ವರೂಪ ಸಮರ್ಥನೀಯವಲ್ಲ. ಸಮರ್ಥನೀಯ: JPEG (.jpg, .jpeg), WebP (.webp)"}, 400

            # Clean up raga and author_name for filename
            raga_clean = raga.replace(" ", "_").replace("/", "_").replace("\\", "_") if raga else "unknown"
            author_clean = author_name.replace(" ", "_").replace("/", "_").replace("\\", "_") if author_name else "unknown"

            # Create filename with naming convention
            new_filename = f"{parva_number}_{sandhi_number}_{padya_number}_{raga_clean}_{author_clean}.{file_ext}"

            # Get absolute path to static folder
            from flask import current_app
            static_folder = current_app.static_folder
            photo_dir = os.path.join(static_folder, 'photos', 'gamakaPhotos')

            # Ensure directory exists
            os.makedirs(photo_dir, exist_ok=True)

            # Full path to save the file
            filepath = os.path.join(photo_dir, new_filename)

            # Save the file
            file.save(filepath)

            # Return relative path for storage in database (relative to static folder)
            relative_path = f"photos/gamakaPhotos/{new_filename}"

            return {
                "photo_path": relative_path,
                "filename": new_filename,
                "message": "Photo uploaded successfully"
            }, 200

        except Exception as e:
            logger.error("Photo upload failed: %s", e)
            return {"error": f"Photo upload failed: {str(e)}"}, 500

    # ---------------------------------------------
    # AUDIO UPLOAD
    # ---------------------------------------------

    def save_gamaka_audio(self, file, parva_number, sandhi_number, padya_number, raga, author_name, **kwargs):
        """
        Save uploaded audio for GamakaVachana (who sang the padya).

        Naming convention: parva_number_sandhi_number_padya_number_raga_authorname.ext
        Example: 1_1_5_Yaman_RamanandaSagara.mp3

        Returns the relative audio path for storage in database.
        """
        try:
            import os
            from werkzeug.utils import secure_filename

            if not file or file.filename == "":
                return {"error": "No file provided"}, 400

            # Validate file type
            allowed_extensions = {'mp3', 'wav', 'ogg', 'flac', 'm4a', 'aac', 'wma'}
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'mp3'

            if file_ext not in allowed_extensions:
                return {"error": f"File type not allowed. Allowed: {', '.join(allowed_extensions)}"}, 400

            # Clean up raga and author_name for filename
            raga_clean = raga.replace(" ", "_").replace("/", "_").replace("\\", "_") if raga else "unknown"
            author_clean = author_name.replace(" ", "_").replace("/", "_").replace("\\", "_") if author_name else "unknown"

            # Create filename with naming convention
            new_filename = f"{parva_number}_{sandhi_number}_{padya_number}_{raga_clean}_{author_clean}.{file_ext}"

            # Get absolute path to static folder
            from flask import current_app
            static_folder = current_app.static_folder
            audio_dir = os.path.join(static_folder, 'audio', 'gamakaAudio')

            # Ensure directory exists
            os.makedirs(audio_dir, exist_ok=True)

            # Full path to save the file
            filepath = os.path.join(audio_dir, new_filename)

            # Save the file
            file.save(filepath)

            # Return relative path for storage in database (relative to static folder)
            relative_path = f"audio/gamakaAudio/{new_filename}"

            return {
                "audio_path": relative_path,
                "filename": new_filename,
                "message": "Audio uploaded successfully"
            }, 200

        except Exception as e:
            logger.error("Audio upload failed: %s", e)
            return {"error": f"Audio upload failed: {str(e)}"}, 500


# -------------------------------------------------------
# GLOBAL SERVICE INSTANCES
# -------------------------------------------------------

parva_service = ParvaService()
sandhi_service = SandhiService()
padya_service = PadyaService()



